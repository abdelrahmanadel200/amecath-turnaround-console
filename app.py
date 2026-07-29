import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

output_path = 'MEA_Master_Intelligence_Workbook.xlsx'

countries = [
    {'Region':'GCC','Country':'Saudi Arabia','Tier':'Tier 1 Major','HD_Patients':31900,'PD_Patients':3200,'Beds':5528,'RegAuthority':'SFDA / NUPCO','AccessMonths':'6-9','Population':36839707},
    {'Region':'GCC','Country':'UAE','Tier':'Tier 1 Major','HD_Patients':7060,'PD_Patients':850,'Beds':2884,'RegAuthority':'MoHAP / DHA / DOH','AccessMonths':'3-6','Population':9516871},
    {'Region':'GCC','Country':'Qatar','Tier':'Tier 1 Major','HD_Patients':941,'PD_Patients':110,'Beds':1191,'RegAuthority':'MoPH','AccessMonths':'3-5','Population':2712499},
    {'Region':'GCC','Country':'Kuwait','Tier':'Tier 1 Major','HD_Patients':4312,'PD_Patients':180,'Beds':1673,'RegAuthority':'Ministry of Health','AccessMonths':'4-6','Population':4962358},
    {'Region':'GCC','Country':'Oman','Tier':'Tier 1 Major','HD_Patients':4650,'PD_Patients':220,'Beds':1500,'RegAuthority':'MoH Oman','AccessMonths':'4-6','Population':4649855},
    {'Region':'GCC','Country':'Bahrain','Tier':'Tier 1 Major','HD_Patients':739,'PD_Patients':95,'Beds':600,'RegAuthority':'NHRA / MOH','AccessMonths':'4-6','Population':1477469},
    {'Region':'East Africa','Country':'Kenya','Tier':'Tier 2 Regional','HD_Patients':11020,'PD_Patients':0,'Beds':2567,'RegAuthority':'PPB','AccessMonths':'3-4','Population':55100586},
    {'Region':'West Africa','Country':'Nigeria','Tier':'Tier 2 Regional','HD_Patients':22380,'PD_Patients':0,'Beds':1706,'RegAuthority':'NAFDAC','AccessMonths':'6-9','Population':223804632},
    {'Region':'West Africa','Country':'Ghana','Tier':'Tier 2 Regional','HD_Patients':5174,'PD_Patients':0,'Beds':2550,'RegAuthority':'FDA Ghana','AccessMonths':'6-9','Population':34494307},
    {'Region':'East Africa','Country':'Tanzania','Tier':'Tier 2 Regional','HD_Patients':8093,'PD_Patients':0,'Beds':1820,'RegAuthority':'TMDA','AccessMonths':'4-6','Population':67438106},
    {'Region':'Southern Africa','Country':'South Africa','Tier':'Tier 1 Major','HD_Patients':8881,'PD_Patients':0,'Beds':2500,'RegAuthority':'SAHPRA','AccessMonths':'6-12','Population':60414495},
    {'Region':'East Africa','Country':'Sudan','Tier':'Tier 1 Major','HD_Patients':7259,'PD_Patients':0,'Beds':1400,'RegAuthority':'NMPB / MoH','AccessMonths':'6-12','Population':48396555},
    {'Region':'North Africa','Country':'Morocco','Tier':'Tier 1 Major','HD_Patients':22443,'PD_Patients':0,'Beds':2200,'RegAuthority':'AMMPS / MoH','AccessMonths':'6-12','Population':37404874},
]

hospitals = [
    ['GCC','Saudi Arabia','King Faisal Specialist Hospital & Research Centre','Specialized Renal Center',1200,'HD','Both','Tier 1 Major'],
    ['GCC','Saudi Arabia','King Salman Center for Kidney Diseases','Specialized Renal Center',78,'HD','Both','Tier 1 Major'],
    ['GCC','Saudi Arabia','King Fahad Medical City','Government',1300,'HD','Both','Tier 1 Major'],
    ['GCC','UAE','Cleveland Clinic Abu Dhabi','Private',360,'HD','Both','Tier 1 Major'],
    ['GCC','UAE','Sheikh Khalifa Medical City','Government',600,'HD','Both','Tier 1 Major'],
    ['GCC','Qatar','Hamad General Hospital','Government',611,'HD','Both','Tier 1 Major'],
    ['GCC','Kuwait','Mubarak Al-Kabeer Hospital','Government',623,'HD','Both','Tier 1 Major'],
    ['GCC','Oman','The Royal Hospital','Government',694,'HD','Both','Tier 1 Major'],
    ['GCC','Bahrain','Salmaniya Medical Complex','Government',1100,'HD','Both','Tier 1 Major'],
    ['East Africa','Kenya','Kenyatta National Hospital','Government',1800,'HD','HD','Tier 1 Major'],
    ['West Africa','Nigeria','Lagos University Teaching Hospital','Government',761,'HD','HD','Tier 1 Major'],
    ['West Africa','Ghana','Korle Bu Teaching Hospital','Government',2000,'HD','HD','Tier 1 Major'],
    ['East Africa','Tanzania','Muhimbili National Hospital','Government',1500,'HD','HD','Tier 1 Major'],
    ['Southern Africa','South Africa','Chris Hani Baragwanath Academic Hospital','Government',3200,'HD','HD','Tier 1 Major'],
    ['East Africa','Sudan','Ahmed Gasim Kidney Center','Specialized Renal Center',0,'HD','HD','Tier 1 Major'],
    ['North Africa','Morocco','Ibn Sina University Hospital','Government',1090,'HD','HD','Tier 1 Major'],
]

# Add placeholder coverage for other country-specific facilities
for c in ['UAE','Qatar','Kuwait','Oman','Bahrain','Kenya','Nigeria','Ghana','Tanzania','South Africa','Sudan','Morocco']:
    if c not in ['Saudi Arabia']:
        hospitals.append([
            next(x['Region'] for x in countries if x['Country']==c),
            c,
            'To Be Verified Renal Facility',
            'Under Evaluation',
            0,
            'HD',
            'To Be Verified',
            'Under Evaluation'
        ])

# Deduplicate and keep a manageable template size
seen = set()
clean_hospitals = []
for row in hospitals:
    key = tuple(row[:3])
    if key in seen:
        continue
    seen.add(key)
    clean_hospitals.append(row)

hospitals = clean_hospitals

distributors = [
    ['Saudi Arabia','Tamer Group','Dialysis / ICU','National','Yes','Target Partner'],
    ['Saudi Arabia','Attieh Medico','ICU / Urology / Dialysis','National','Yes','Target Partner'],
    ['UAE','Zahrawi Group','Dialysis / ICU','Regional','Yes','Target Partner'],
    ['UAE','GulfDrug','Dialysis Equipment / Consumables','National','Yes','Target Partner'],
    ['Kuwait','Yiaco Medical','Medical Devices / Dialysis','National','Yes','Target Partner'],
    ['Qatar','Al Muftah Co Healthcare Division','General Medical Equipment','National','Yes','Prospect'],
    ['Kenya','Crown Healthcare','General Medical Equipment','Regional','Yes','Target Partner'],
    ['Nigeria','JNC International Limited','ICU / Imaging / Turnkey','Regional','Yes','Target Partner'],
    ['Ghana','Ibermansa Ghana','Equipment Planning / Procurement','National','Yes','Target Partner'],
    ['Tanzania','Africa Healthcare Network','Dialysis Services / Consumables','Regional','Yes','Target Partner'],
    ['South Africa','To Be Verified','Dialysis / Renal','National','Under Evaluation','Under Evaluation'],
    ['Sudan','To Be Verified','Dialysis / Renal','National','Under Evaluation','Under Evaluation'],
    ['Morocco','To Be Verified','Dialysis / Renal','National','Under Evaluation','Under Evaluation'],
    ['Oman','To Be Verified','Dialysis / Renal','National','Under Evaluation','Under Evaluation'],
    ['Bahrain','To Be Verified','Dialysis / Renal','National','Under Evaluation','Under Evaluation'],
]

kols = [
    ['Nephrology / Dialysis Unit','Head of Nephrology / Dialysis Medical Director','HD vascular access longevity, infection rates, catheter dwell time','Clinical evidence for lower infection and improved patency','Acute HD / Permcath'],
    ['Interventional Nephrology','Interventional Nephrologist','First-pass insertion success, access preservation','Reliable insertion performance and bedside usability','Acute HD / Permcath'],
    ['Procurement / Supply Chain','Procurement Officer / Value Analysis Chair','Budget control and tender compliance','Competitive pricing with compliant documentation','Acute HD / Permcath / PD'],
    ['Peritoneal Dialysis Program','Nephrology Program Lead','PD catheter migration, leaks, exit-site infection','PD-focused access solutions for home program expansion','PD Catheters'],
]

regulatory = [
    ['Saudi Arabia','SFDA / NUPCO','MDMA via GHAD portal; local authorized rep required','CE / ISO 13485 + SFDA technical file','6-9'],
    ['UAE','MoHAP / DHA / DOH','Medical device registration via federal/emirate pathways','CE / ISO 13485','3-6'],
    ['Qatar','MoPH','Local distributor submission; FSC required','CE / ISO 13485','3-5'],
    ['Kuwait','Ministry of Health','Local agent registration; FSC required','CE / ISO 13485','4-6'],
    ['Oman','MoH Oman','Local agent/distributor registration','CE / ISO 13485','4-6'],
    ['Bahrain','NHRA / MOH','Local distributor registration','CE / ISO 13485','4-6'],
    ['Kenya','PPB','Authorized representative / registered distributor','CE / ISO 13485','3-4'],
    ['Nigeria','NAFDAC','Local representative + NAFDAC registration','CE / ISO 13485','6-9'],
    ['Ghana','FDA Ghana','Local registration and import clearance','CE / ISO 13485','6-9'],
    ['Tanzania','TMDA','Device registration and importer listing','CE / ISO 13485','4-6'],
    ['South Africa','SAHPRA','Risk-based registration; local rep required','CE / ISO 13485','6-12'],
    ['Sudan','NMPB / MoH','Local registration pathway','CE / ISO 13485','6-12'],
    ['Morocco','AMMPS / MoH','Local registration pathway','CE / ISO 13485','6-12'],
]

blended_price_hd = 85
blended_price_pd = 120
hd_ratio_of_beds = 0.04
sessions_per_week = 3
hd_catheters_per_100_sessions = 2
pd_catheters_per_patient_year = 2
gcc_pd_penetration = 0.12

def is_gcc(region):
    return region == 'GCC'

# Build financial model
fm_rows = []
for c in countries:
    active_pd = c['PD_Patients'] if is_gcc(c['Region']) else 0
    annual_hd_units = round(c['HD_Patients'] * hd_catheters_per_100_sessions / 100.0, 2)
    annual_pd_units = round(active_pd * pd_catheters_per_patient_year, 2) if is_gcc(c['Region']) else 0
    revenue = round(annual_hd_units * blended_price_hd + annual_pd_units * blended_price_pd, 2)
    fm_rows.append([
        c['Region'], c['Country'], c['Tier'], c['HD_Patients'], active_pd,
        annual_hd_units, annual_pd_units, f'{blended_price_hd}/{blended_price_pd}', revenue
    ])

wb = Workbook()
ws = wb.active
ws.title = 'Sheet 1_Target Hospitals'

headers1 = ['Region','Country','Facility Name','Sector','Estimated Bed Count','Nephrology & Dialysis Capabilities (HD / PD)','Target AMECATH Dialysis Portfolio (HD / PD / Both)','Account Tier']
ws.append(headers1)
for row in hospitals:
    ws.append(row)

ws2 = wb.create_sheet('Sheet 2_Distributors')
headers2 = ['Country','Distributor Name','Specialty Focus','Market Coverage','Tendering Capability','Potential Pipeline Status']
ws2.append(headers2)
for row in distributors:
    ws2.append(row)

ws3 = wb.create_sheet('Sheet 3_KOLs')
headers3 = ['Department / Specialty','Key Decision Maker Title / Role','Clinical Pain Point / Priority','AMECATH Value Proposition','Target Product Focus']
ws3.append(headers3)
for row in kols:
    ws3.append(row)

ws4 = wb.create_sheet('Sheet 4_Regulatory')
headers4 = ['Country','Target Authority','Required Certificates / Registration Route','AMECATH Compliance Route (CE / ISO / SFDA)','Estimated Time-to-Market (Months)']
ws4.append(headers4)
for row in regulatory:
    ws4.append(row)

ws5 = wb.create_sheet('Sheet 5_Financial Model')
headers5 = ['Region','Country','Strategic Tier','Active Dialysis Patients (HD)','Active Dialysis Patients (PD - GCC Focused)','Estimated Annual HD Catheter Units Needed','Estimated Annual PD Catheter Units Needed','Blended Unit Price (USD)','Projected Annual Revenue Potential (USD)']
ws5.append(headers5)
for row in fm_rows:
    ws5.append(row)

ws6 = wb.create_sheet('References')
refs = [
    ['Source Type','Reference / Notes','URL / Identifier'],
    ['Workbook baseline','AMECATH_Market_Intelligence-1.xlsx','Local file'],
    ['ISN-GKHA','Middle East narrative script and regional kidney health context','https://www.theisn.org/wp-content/uploads/2024/01/ISN-GKHA-2023_Narrative_Script_Middle-East_v0.1.pdf'],
    ['South Africa registry','South African Renal Registry annual reports 2022 and 2023','https://www.sahr.org.za/'],
    ['Saudi Arabia regulation','SFDA import permits / medical device registration','https://www.sfda.gov.sa/'],
    ['UAE regulation','MOHAP medical device and health facility regulation','https://mohap.gov.ae/'],
    ['Qatar regulation','Ministry of Public Health medical device registration','https://www.moph.gov.qa/'],
    ['Kuwait regulation','Ministry of Health device registration','https://www.moh.gov.kw/'],
    ['Kenya regulation','Pharmacy and Poisons Board','https://www.pharmacyboardkenya.org/'],
    ['Nigeria regulation','NAFDAC medical devices / import control','https://www.nafdac.gov.ng/'],
    ['Ghana regulation','Food and Drugs Authority','https://fdaghana.gov.gh/'],
    ['Tanzania regulation','Tanzania Medicines and Medical Devices Authority','https://www.tmda.go.tz/'],
    ['South Africa regulation','SAHPRA medical device licensing','https://www.sahpra.org.za/'],
]
for row in refs:
    ws6.append(row)

# Formatting
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
for wsx in wb.worksheets:
    for cell in wsx[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wsx.freeze_panes = 'A2'
    for col in wsx.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = len(str(cell.value)) if cell.value is not None else 0
                if val > max_len:
                    max_len = val
            except:
                pass
        wsx.column_dimensions[col_letter].width = min(max_len + 2, 45)

wb.save(output_path)
print(output_path)
